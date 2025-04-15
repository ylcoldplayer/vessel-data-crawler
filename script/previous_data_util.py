import pandas as pd


def read_excl_file():
    # Path to your Excel file
    excel_file_path = '../data/市场数据跟踪 二次调整颜色(1).xlsx'

    # Load the Excel file
    xl = pd.ExcelFile(excel_file_path)

    # Print the sheet names
    print("Sheet Names:", xl.sheet_names)

    # Display contents of each sheet
    for sheet_name in xl.sheet_names:
        print(f"\nDisplaying Sheet: {sheet_name}")
        df = xl.parse(sheet_name)
        print(df)



import openpyxl

def display_all_columns(excel_file_path):
    # Load the workbook and sheet names
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        print(f"\nDisplaying Sheet: {sheet_name}")
        sheet = workbook[sheet_name]

        # Get max column for iterating through columns
        max_column = sheet.max_column

        # Gather all column headers, including hidden ones
        columns = []
        for col_idx in range(1, max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            col_dim = sheet.column_dimensions[col_letter]

            # Check if the column is hidden
            if col_dim.hidden:
                # Handle hidden columns
                columns.append(f"Hidden Column {col_letter}")
            else:
                cell_value = sheet.cell(row=1, column=col_idx).value
                columns.append(cell_value if cell_value is not None else f"Column {col_letter}")

        # Display column headers
        print("Columns:", columns)

        # Display rows (starting from row 2 assuming row 1 has headers)
        for row in sheet.iter_rows(min_row=2):
            row_values = [cell.value for cell in row]
            print(row_values)


def read_excel_to_dataframe(excel_file_path):
    # Load the first sheet of the Excel file into a DataFrame
    df = pd.read_excel(excel_file_path, sheet_name='日报')

    # Return the DataFrame
    return df


def read_shanghai_10mm():
    excel_file_path = '../data/市场数据跟踪 二次调整颜色(1).xlsx'
    df = read_excel_to_dataframe(excel_file_path)
    # print(df)
    df_shanghai_10mm = df[df['Daily Market Data\n每日市场数据'] == '上海10mm船板价格'].iloc[:, :215]
    # print(df_shanghai_10mm)
    df_date = df.iloc[[1]].reset_index(drop=True).iloc[:, :215]

    # print(df_date)

    result_df = pd.concat([df_date, df_shanghai_10mm], ignore_index=True)
    print(result_df)
    return result_df


def read_row_data(row_name='上海10mm船板价格'):
    excel_file_path = '../data/市场数据跟踪 二次调整颜色(1).xlsx'

    df = read_excel_to_dataframe(excel_file_path)
    # print(df)
    df_shanghai_10mm = df[df['Daily Market Data\n每日市场数据'] == row_name].iloc[:, :215]
    # print(df_shanghai_10mm)
    df_date = df.iloc[[1]].reset_index(drop=True).iloc[:, :215]

    # print(df_date)

    result_df = pd.concat([df_date, df_shanghai_10mm], ignore_index=True)
    print(result_df)
    return result_df


if __name__ == '__main__':
    column_list = ['IFO 380cst Bunker Prices (3.5% Sulphur), Singapore $/Tonne', 'CCFI\n中国出口集装箱运价指数','BDTI\n波罗的海原油运价指数','BCTI\n波罗的海成品油运价指数', 'BDI\n波罗的海干散货运价指数' ,'上海10mm船板价格', '上海20mm船板价格', '广州10mm船板价格', '广州20mm船板价格', '南京10mm船板价格', '南京20mm船板价格']

    df = read_row_data(column_list[0])
    print(df.columns[1:])
    for col in df.columns[1:]:
        print(df[col][0])
        print(df[col][1])
